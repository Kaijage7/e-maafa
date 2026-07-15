#!/usr/bin/env python3
"""Export local-only DMIS incident-flow login workbooks from the current database.

The exporter never reads password hashes. Every listed account is reset to the documented
local test password by LocalTestPasswordSeeder whenever the backend starts with the `local`
Spring profile. The generated credential files are ignored by Git.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
from pathlib import Path
import subprocess
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


LOCAL_PASSWORD = "Password@2026"
LOGIN_URL = "http://localhost:4200/login"
ROLE_ORDER = {
    "Dist DC": (1, "Stage owner", "waiting_ddmc", "Approve/escalate or close as rumour within own area"),
    "DED": (2, "Stage owner", "waiting_ded", "Approve/escalate within own district/council area"),
    "DAS": (3, "District adviser", "View/comment", "Advisory comments; does not approve incidents"),
    "District Commissioner": (4, "District adviser", "View/comment", "Area oversight and advisory comments"),
    "District Planning Officer": (5, "District adviser", "View/comment", "Planning advice and incident visibility"),
    "District Logistic Officer": (6, "District logistics", "Logistics", "Warehouse, allocation, dispatch and incident visibility"),
    "Reg DC": (7, "Stage owner", "waiting_rdmc", "Regional coordinator approval/escalation"),
    "RAS": (8, "Stage owner", "waiting_ras", "Regional administrative approval/escalation"),
    "RC": (9, "Regional adviser", "View/comment", "Regional oversight and advisory comments"),
    "Regional Planning Officer": (10, "Regional adviser", "View/comment", "Regional planning advice and visibility"),
    "Regional Logistic Officer": (11, "Regional logistics", "Logistics", "Regional warehouse, allocation and dispatch"),
    "EOCC": (12, "National stage owner", "waiting_eocc", "EOCC operational approval and coordination"),
    "Director": (13, "National stage owner", "waiting_director", "Director approval/escalation"),
    "Secretary": (14, "National stage owner", "waiting_ps", "Principal Secretary final approval"),
}
ROLE_SQL_LIST = ", ".join("'" + name.replace("'", "''") + "'" for name in ROLE_ORDER)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--host", default=os.getenv("DB_HOST", "127.0.0.1"))
    parser.add_argument("--port", default=os.getenv("DMIS_DB_PUBLISH_PORT", "5440"))
    parser.add_argument("--database", default=os.getenv("DB_NAME", "dmis"))
    parser.add_argument("--user", default=os.getenv("DB_USERNAME", "dmis_app"))
    parser.add_argument("--output-dir", type=Path, default=Path("test-data/local"))
    return parser.parse_args()


def query_csv(args: argparse.Namespace, sql: str) -> list[dict[str, str]]:
    env = os.environ.copy()
    env.setdefault("PGPASSWORD", os.getenv("DB_PASSWORD", "dmis_pass"))
    command = [
        "psql", "-X", "--csv", "-q", "-v", "ON_ERROR_STOP=1",
        "-h", args.host, "-p", str(args.port), "-U", args.user, "-d", args.database,
        "-c", sql,
    ]
    completed = subprocess.run(command, env=env, text=True, capture_output=True)
    if completed.returncode != 0:
        detail = completed.stderr.strip() or "psql returned no diagnostic"
        raise SystemExit(f"Credential export database query failed: {detail}")
    return list(csv.DictReader(io.StringIO(completed.stdout)))


def ddmc_coverage(args: argparse.Namespace) -> list[dict[str, str]]:
    return query_csv(args, """
        with ddmc as (
            select distinct u.*
              from public.users u
              join public.model_has_roles mhr
                on mhr.model_id = u.id and mhr.model_type = 'App\\Models\\User'
              join public.roles role on role.id = mhr.role_id and role.name = 'Dist DC'
        )
        select c.id::text as council_id,
               c.council_code,
               c.name as council,
               c.country_part,
               d.id::text as district_id,
               d.name as district,
               region.id::text as region_id,
               region.name as region,
               chosen.id::text as user_id,
               chosen.name as user_name,
               chosen.email,
               chosen.position_key,
               case when chosen.council_id = c.id then 'council seat'
                    else 'Zanzibar district seat mapped to council' end as coverage_source
          from public.councils c
          join public.districts d on d.id = c.district_id
          join public.regions region on region.id = c.region_id
          left join lateral (
              select u.id, u.name, u.email, u.position_key, u.council_id
                from ddmc u
               where u.council_id = c.id
                  or (c.country_part = 'zanzibar'
                      and u.council_id is null
                      and u.district_id = c.district_id
                      and coalesce(u.seeded_officer, false))
               order by case when u.council_id = c.id then 0 else 1 end, u.id
               limit 1
          ) chosen on true
         where coalesce(c.is_active, true)
         order by region.name, d.name, c.name
    """)


def all_incident_roles(args: argparse.Namespace) -> list[dict[str, str]]:
    return query_csv(args, f"""
        select distinct u.id::text as user_id,
               role.name as role,
               u.name as user_name,
               u.email,
               coalesce(region.country_part, district.country_part, council.country_part, 'national') as country_part,
               coalesce(region.name, '') as region,
               coalesce(district.name, '') as district,
               coalesce(council.name, '') as council,
               coalesce(u.region_id::text, '') as region_id,
               coalesce(u.district_id::text, '') as district_id,
               coalesce(u.council_id::text, '') as council_id,
               coalesce(u.position_key, '') as position_key,
               case when coalesce(u.seeded_officer, false) then 'seeded position' else 'named account' end as account_kind,
               case when u.password is not null and not coalesce(u.must_change_password, false)
                    then 'ready before local restart' else 'reset on next local start' end as current_state
          from public.users u
          join public.model_has_roles mhr
            on mhr.model_id = u.id and mhr.model_type = 'App\\Models\\User'
          join public.roles role on role.id = mhr.role_id
          left join public.regions region on region.id = u.region_id
          left join public.districts district on district.id = u.district_id
          left join public.councils council on council.id = u.council_id
         where role.name in ({ROLE_SQL_LIST})
         order by role, region nulls first, district nulls first,
                  council nulls first, email
    """)


def style_data_sheet(ws, title: str, warning: str, headers: list[str], rows: Iterable[list[object]], table_name: str) -> int:
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A2"] = warning
    ws["A2"].font = Font(bold=True, color="9C0006")
    ws["A2"].fill = PatternFill("solid", fgColor="FFC7CE")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    for column, header in enumerate(headers, 1):
        cell = ws.cell(3, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    count = 0
    for count, row in enumerate(rows, 1):
        for column, value in enumerate(row, 1):
            ws.cell(count + 3, column, value)
    last_row = max(3, count + 3)
    if count:
        table = Table(displayName=table_name, ref=f"A3:{ws.cell(last_row, len(headers)).coordinate}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(table)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{ws.cell(last_row, len(headers)).coordinate}"
    for column_index, column_cells in enumerate(ws.iter_cols(), 1):
        letter = get_column_letter(column_index)
        width = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 55)
    return count


def write_csv(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def read_optional_report(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        report = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return report if isinstance(report, dict) else {}


def report_summary(report: dict, expected_total: int) -> str:
    if not report:
        return "NOT RUN"
    total = report.get("total")
    passed = report.get("passed")
    failed = report.get("failed")
    suffix = "" if total == expected_total else f" (expected {expected_total})"
    return f"{passed}/{total} PASS; {failed} FAIL{suffix}"


def transition_pass(result: dict, stage: str, allowed_targets: set[str]) -> bool:
    return any(
        transition.get("from") == stage
        and transition.get("http_status") == 200
        and transition.get("actual_to") in allowed_targets
        for transition in result.get("stage_transitions", [])
    )


def main() -> None:
    args = parse_args()
    ddmc = ddmc_coverage(args)
    all_roles = all_incident_roles(args)

    auth_report = read_optional_report(args.output_dir / "incident-persona-auth-results.json")
    scope_report = read_optional_report(args.output_dir / "ddmc-195-scope-results.json")
    workflow_report = read_optional_report(args.output_dir / "ddmc-195-workflow-results.json")
    scope_by_test = {str(row.get("test_number")): row for row in scope_report.get("results", [])}
    workflow_by_test = {str(row.get("test_number")): row for row in workflow_report.get("results", [])}

    if len(ddmc) != 195:
        raise SystemExit(f"Expected 195 active councils, database returned {len(ddmc)}")
    if any(not row["user_id"] or not row["email"] for row in ddmc):
        missing = [row["council"] for row in ddmc if not row["user_id"] or not row["email"]]
        raise SystemExit(f"DDMC coverage is incomplete: {missing}")
    if len({row["council_id"] for row in ddmc}) != 195 or len({row["email"] for row in ddmc}) != 195:
        raise SystemExit("DDMC export must contain 195 unique councils and 195 unique logins")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = args.output_dir / "DMIS_LOCAL_INCIDENT_FLOW_CREDENTIALS.xlsx"
    ddmc_csv_path = args.output_dir / "DMIS_LOCAL_DDMC_195.csv"
    all_csv_path = args.output_dir / "DMIS_LOCAL_ALL_INCIDENT_PERSONAS.csv"

    warning = "LOCAL TESTING ONLY — never use these shared credentials under prod or on a public edge."
    ddmc_headers = [
        "Test #", "Country part", "Region", "District", "Council/LGA", "Council code",
        "Coverage source", "Name", "Email", "Password", "Role", "Workflow stage",
        "Login URL", "User ID", "Region ID", "District ID", "Council ID", "Position key",
    ]
    ddmc_rows = [
        [
            index, row["country_part"], row["region"], row["district"], row["council"],
            row["council_code"], row["coverage_source"], row["user_name"], row["email"],
            LOCAL_PASSWORD, "Dist DC", "waiting_ddmc", LOGIN_URL, row["user_id"],
            row["region_id"], row["district_id"], row["council_id"], row["position_key"],
        ]
        for index, row in enumerate(ddmc, 1)
    ]

    all_headers = [
        "Test #", "Flow order", "Role class", "Role", "Stage/action", "Expected use",
        "Country part", "Region", "District", "Council/LGA", "Name", "Email", "Password",
        "Login URL", "Account kind", "Current DB state", "User ID", "Region ID", "District ID",
        "Council ID", "Position key",
    ]
    all_roles.sort(key=lambda row: (ROLE_ORDER[row["role"]][0], row["region"], row["district"], row["council"], row["email"]))
    all_rows = []
    for index, row in enumerate(all_roles, 1):
        order, role_class, stage, use = ROLE_ORDER[row["role"]]
        all_rows.append([
            index, order, role_class, row["role"], stage, use, row["country_part"], row["region"],
            row["district"], row["council"], row["user_name"], row["email"], LOCAL_PASSWORD,
            LOGIN_URL, row["account_kind"], row["current_state"], row["user_id"], row["region_id"],
            row["district_id"], row["council_id"], row["position_key"],
        ])

    wb = Workbook()
    readme = wb.active
    readme.title = "READ ME"
    readme.sheet_view.showGridLines = False
    readme["A1"] = "DMIS local incident-flow credential pack"
    readme["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    readme["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    readme.merge_cells("A1:D1")
    notes = [
        ("Safety", warning),
        ("Local password", LOCAL_PASSWORD),
        ("Login URL", LOGIN_URL),
        ("DDMC council coverage", f"{len(ddmc)} of 195 active councils/LGAs"),
        ("All incident persona rows", str(len(all_rows))),
        ("Latest login/JWT sweep", report_summary(auth_report, 1309)),
        ("Latest DDMC scope sweep", report_summary(scope_report, 195)),
        ("Latest DDMC lifecycle sweep", report_summary(workflow_report, 195)),
        ("Latest controlled cleanup", str(workflow_report.get("cleanup") or "NOT RUN")),
        ("Execution tracker", "DDMC test tracker contains one row per active council with dropdown results for the full incident ladder, negative scope checks, UI/API checks and cleanup."),
        ("How credentials become valid", "Start the backend with Spring profile local. LocalTestPasswordSeeder resets every listed role to the local password and clears must_change_password."),
        ("Production rule", "Never run the local profile publicly. Production migrations revoke demo hashes; real operators need unique passwords and 2FA."),
        ("Zanzibar mapping", "The schema has 11 active Zanzibar councils and 11 district DDMC positions, each attached to its one-to-one council. The unstaffed RDMC tier auto-skips to the real regional RAS and is recorded in workflow history."),
    ]
    for row_index, (label, value) in enumerate(notes, 3):
        readme.cell(row_index, 1, label).font = Font(bold=True)
        readme.cell(row_index, 2, value).alignment = Alignment(wrap_text=True, vertical="top")
    readme.column_dimensions["A"].width = 30
    readme.column_dimensions["B"].width = 110

    ddmc_ws = wb.create_sheet("DDMC 195")
    style_data_sheet(ddmc_ws, "DDMC credentials — all 195 active councils/LGAs", warning,
                     ddmc_headers, ddmc_rows, "DDMC195Credentials")

    tracker_headers = [
        "Test #", "Country part", "Region", "District", "Council/LGA", "DDMC email",
        "Incident/report reference", "Create/convert", "Own-scope visibility", "DDMC decision",
        "DED decision", "RDMC decision", "RAS decision", "EOCC decision", "Director decision",
        "PS decision", "Cross-area denial", "Adviser comment", "Logistics checks", "UI/browser",
        "API", "Cleanup", "Overall result", "Tester", "Test date", "Notes",
    ]
    tracker_rows = []
    cleanup_ok = (
        workflow_report.get("cleanup", {}).get("incidents_after") == 0
        and workflow_report.get("cleanup", {}).get("notifications_after") == 0
    )
    workflow_marker = str(workflow_report.get("marker") or "")
    test_date = str(workflow_report.get("started_at") or "")[:10]
    for index, row in enumerate(ddmc, 1):
        key = str(index)
        scope_result = scope_by_test.get(key, {})
        workflow_result = workflow_by_test.get(key, {})
        if not workflow_result:
            statuses = ["NOT RUN"] * 16
            reference = ""
            notes_value = ""
        else:
            issues = workflow_result.get("issues") or []
            scope_issues = scope_result.get("issues") or []
            zanzibar_rdmc_skip = "waiting_rdmc" in workflow_result.get("auto_skipped_stages", [])
            statuses = [
                "PASS" if workflow_result.get("create_status") == 200 else "FAIL",
                "PASS" if workflow_result.get("own_show_status") == 200 and not scope_issues else "FAIL",
                "PASS" if workflow_result.get("submit_stage") == "waiting_ddmc"
                          and transition_pass(workflow_result, "waiting_ddmc", {"waiting_ded"}) else "FAIL",
                "PASS" if workflow_result.get("adviser_approve_status") == 403
                          and transition_pass(workflow_result, "waiting_ded", {"waiting_rdmc", "waiting_ras"}) else "FAIL",
                "N/A" if zanzibar_rdmc_skip else (
                    "PASS" if transition_pass(workflow_result, "waiting_rdmc", {"waiting_ras"}) else "FAIL"
                ),
                "PASS" if transition_pass(workflow_result, "waiting_ras", {"waiting_eocc"}) else "FAIL",
                "PASS" if transition_pass(workflow_result, "waiting_eocc", {"waiting_director"}) else "FAIL",
                "PASS" if transition_pass(workflow_result, "waiting_director", {"waiting_ps"}) else "FAIL",
                "PASS" if transition_pass(workflow_result, "waiting_ps", {"approved"}) else "FAIL",
                "PASS" if workflow_result.get("foreign_show_status") == 404 else "FAIL",
                "NOT RUN",  # adviser approval denial passed; comment creation is still a separate scenario
                "NOT RUN",
                "NOT RUN",
                "PASS" if not issues and not scope_issues else "FAIL",
                "PASS" if cleanup_ok else "FAIL",
                "PASS" if not issues and not scope_issues and cleanup_ok else "FAIL",
            ]
            incident_id = workflow_result.get("incident_id")
            reference = f"{workflow_marker}; incident {incident_id} (cleaned)"
            notes = ["Automated API lifecycle and scope sweep; manual incident creation tested, public-report conversion not tested."]
            if zanzibar_rdmc_skip:
                notes.append("RDMC is unstaffed by design; verified auto_advanced history to staffed RAS.")
            notes.append("Adviser comment, logistics operations and browser UI remain NOT RUN.")
            notes_value = " ".join(notes)
        tracker_rows.append([
            index, row["country_part"], row["region"], row["district"], row["council"], row["email"],
            reference, *statuses, "Codex local automated harness" if workflow_result else "", test_date if workflow_result else "", notes_value,
        ])
    tracker_ws = wb.create_sheet("DDMC test tracker")
    style_data_sheet(tracker_ws, "Incident-flow execution tracker — all 195 councils/LGAs", warning,
                     tracker_headers, tracker_rows, "DDMC195TestTracker")
    result_validation = DataValidation(
        type="list", formula1='"NOT RUN,PASS,FAIL,BLOCKED,N/A"', allow_blank=False
    )
    result_validation.error = "Choose NOT RUN, PASS, FAIL, BLOCKED or N/A"
    result_validation.errorTitle = "Invalid test result"
    tracker_ws.add_data_validation(result_validation)
    for column_index in range(8, 24):
        letter = get_column_letter(column_index)
        result_validation.add(f"{letter}4:{letter}{len(tracker_rows) + 3}")

    all_ws = wb.create_sheet("All incident personas")
    style_data_sheet(all_ws, "Maximum local incident-flow persona coverage", warning,
                     all_headers, all_rows, "AllIncidentPersonas")

    summary_headers = ["Flow order", "Role", "Role class", "Stage/action", "Expected use", "Accounts"]
    summary_rows = []
    for role, (order, role_class, stage, use) in sorted(ROLE_ORDER.items(), key=lambda item: item[1][0]):
        summary_rows.append([order, role, role_class, stage, use, sum(1 for row in all_roles if row["role"] == role)])
    summary_ws = wb.create_sheet("Role summary")
    style_data_sheet(summary_ws, "Incident flow and persona counts", warning,
                     summary_headers, summary_rows, "IncidentRoleSummary")

    scenario_headers = ["#", "Scenario", "Primary persona", "Expected result", "Evidence to capture"]
    scenario_rows = [
        [1, "Local login", "Any listed persona", "Login succeeds only after local-profile seeder runs", "HTTP status, role, area payload and screenshot"],
        [2, "Own-area report conversion", "Dist DC", "Report converts to own-area incident at the DDMC/DED entry path", "Report code, incident ID and workflow history"],
        [3, "Own-area manual incident", "Dist DC", "Own council/district target is accepted", "Incident ID and stored region/district/council"],
        [4, "Foreign-area create/update", "Dist DC", "Foreign district/council and assignee are rejected", "422/404 response and zero foreign row mutation"],
        [5, "DDMC stage decision", "Dist DC", "Only own-area waiting_ddmc incident can be advanced or closed as rumour", "Before/after state and history actor"],
        [6, "DED stage decision", "DED", "Own-area waiting_ded advances; DDMC and advisers cannot approve it", "Positive and negative HTTP results"],
        [7, "Regional coordinator decision", "Reg DC", "Own-region waiting_rdmc advances; foreign region is hidden/denied", "Workflow history and scope denial"],
        [8, "RAS decision", "RAS", "Own-region waiting_ras advances; other regions remain isolated", "Workflow history and negative probe"],
        [9, "National approval chain", "EOCC / Director / Secretary", "Each national stage owner acts only at its assigned stage", "waiting_eocc → waiting_director → waiting_ps → approved"],
        [10, "Advisory separation", "DAS / DC / planning officers / RC", "Can view/comment in scope but cannot approve", "Comment row plus 403 approval attempt"],
        [11, "Logistics separation", "District/Regional Logistic Officer", "Can use scoped stock/dispatch functions but cannot approve incidents", "Allocation/dispatch evidence plus 403 approval"],
        [12, "Cross-area list isolation", "District and regional personas", "Lists, counts, form data and detail exclude foreign areas", "API totals and controlled marker searches"],
        [13, "Browser action visibility", "Every role class", "UI hides actions not granted while backend independently denies them", "Screenshot plus direct forbidden API call"],
        [14, "Notification safety", "Stage owners", "In-app workflow notification is created; external SMS/email remains disabled for local bulk testing", "Notification row and zero external delivery"],
        [15, "Idempotence and cleanup", "Tester/admin", "Repeat decisions do not double-advance; controlled data is removed", "History counts and zero marker leftovers"],
    ]
    scenario_ws = wb.create_sheet("Scenario catalogue")
    style_data_sheet(scenario_ws, "Maximum local incident-flow scenario catalogue", warning,
                     scenario_headers, scenario_rows, "IncidentScenarioCatalogue")

    wb.save(workbook_path)
    write_csv(ddmc_csv_path, ddmc_headers, ddmc_rows)
    write_csv(all_csv_path, all_headers, all_rows)
    print(f"Generated {workbook_path} with {len(ddmc_rows)} DDMC rows and {len(all_rows)} total persona rows")
    print(f"Generated {ddmc_csv_path}")
    print(f"Generated {all_csv_path}")


if __name__ == "__main__":
    main()
